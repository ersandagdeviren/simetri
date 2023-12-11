from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import portrait, A4
from reportlab.lib.units import cm
from openpyxl import load_workbook
import textwrap
import qrcode
import os
import tkinter as tk
import pandas as pd

# Load data from Excel file
wb = load_workbook('/Users/ersandagdeviren/Desktop/toners.xlsx')
ws = wb.active

# Define label dimensions and margins
label_width = 9.91 * cm
label_height = 3.35 * cm
horizontal_margin = 0.5 * cm
vertical_margin = 1.3 * cm

# Calculate number of rows and columns
num_rows = ws.max_row
num_cols = ws.max_column

# Calculate number of labels per page
labels_per_row = 2
labels_per_col = 8
labels_per_page = labels_per_row * labels_per_col

# Create an empty list for labels and qrcodes
labels_list = []
qrcode_list = []
title_list=[]

# Function to add label and qrcode to the lists
def add_label():
    product_code = product_code_entry.get()
    num_labels = int(number_of_labels_entry.get())
    excel_path = "/Users/ersandagdeviren/Desktop/toners.xlsx"

    try:
        df = pd.read_excel(excel_path)
        filtered_df = df[df['stockCode'].str.contains(product_code, case=False, na=False)]

        if not filtered_df.empty:
            label = filtered_df.iloc[0]['label']
            qrcode = filtered_df.iloc[0]['qrCode']
            title=filtered_df.iloc[0]['stockCode']
            labels_list.extend([label] * num_labels)
            qrcode_list.extend([qrcode] * num_labels)
            title_list.extend([title] * num_labels)

            # Modify the output text to display the label value
            output_text.config(state=tk.NORMAL)
            output_text.insert(tk.END, f"{num_labels} number of labels for {label} is added\n")
            output_text.config(state=tk.DISABLED)
        else:
            output_text.config(state=tk.NORMAL)
            output_text.insert(tk.END, f"No matching product found for code: {product_code}\n")
            output_text.config(state=tk.DISABLED)

    except Exception as e:
        output_text.config(state=tk.NORMAL)
        output_text.insert(tk.END, f"Error: {str(e)}\n")
        output_text.config(state=tk.DISABLED)

def clear_labels():
    global labels_list, qrcode_list, title_list
    labels_list = []
    qrcode_list = []
    title_list=[]
    output_text.config(state=tk.NORMAL)
    output_text.delete(1.0, tk.END)
    output_text.config(state=tk.DISABLED)

def execute_label():
    # Check if there are labels and qrcodes in the lists
    if labels_list and qrcode_list and title_list:
        # Create a PDF file to store the labels
        if not os.path.exists('/Users/ersandagdeviren/Desktop/'):
            os.makedirs('/Users/ersandagdeviren/Desktop/')

        pdf = canvas.Canvas('/Users/ersandagdeviren/Desktop/executed_labels.pdf', pagesize=portrait(A4))

        # Loop through the labels and qrcodes and generate labels
        for i in range(len(labels_list)):
            product_title=title_list[i]                       #product_title = labels_list[i]
            qr_text = qrcode_list[i]
            label_text=labels_list[i]   #added

            # Wrap product details text (you can customize this part based on your requirements)
            product_details = label_text

            # Generate QR code
            qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
            qr.add_data(qr_text)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            img_path = f"qr_executed_{i}.png"
            img.save(img_path)

            # Calculate label position
            row_index = (i % labels_per_page) // labels_per_row
            col_index = i % labels_per_row
            x = horizontal_margin + col_index * label_width
            y = A4[1] - vertical_margin - (row_index + 1) * label_height

            # Set the color of the stroke to white
            pdf.setStrokeColorRGB(1, 1, 1)  # 1,1,1 corresponds to white

            # Draw label rectangle
            pdf.rect(x, y, label_width, label_height)

            # Reset the stroke color to black (if necessary)
            pdf.setStrokeColorRGB(0, 0, 0)  # 0,0,0 corresponds to black

            # Draw product title
            pdf.setFont('Helvetica-Bold', 16)
            #pdf.drawString(x + 5, y + label_height - 20, " " + product_title + "  For Use In ")
            line_height = 12
            max_lines = int((label_height - 70) / line_height)
            wrapped_text = textwrap.wrap(product_title, width=50)#dfsfsdfsdfssfsfsdf
            num_lines = min(len(wrapped_text), max_lines)

            for j in range(num_lines):
                pdf.drawString(x + 5, y + label_height - 20 ," " + wrapped_text[j]+" For Use In")


            # Draw line between title and description
            pdf.line(x + 5, y + label_height - 30, x + label_width - 5, y + label_height - 30)

            # Draw product details
            pdf.setFont('Helvetica', 10)
            line_height = 12
            max_lines = int((label_height - 70) / line_height)
            wrapped_text = textwrap.wrap(product_details, width=50)#dfsfsdfsdfssfsfsdf
            num_lines = min(len(wrapped_text), max_lines)

            for j in range(num_lines):
                pdf.drawString(x + 5, y + label_height - 40 - j * line_height, " " + wrapped_text[j])

            # Draw QR code 
            pdf.drawImage(img_path, x + label_width - 1.5 * cm - 5, y + 5.15, width=1.5 * cm, height=1.5 * cm)

            # Add page break if necessary
            if (i + 1) % labels_per_page == 0:
                pdf.showPage()

        # Save the generated PDF file
        pdf.save()

        # Notify the user that label generation is complete
        output_text.config(state=tk.NORMAL)
        output_text.insert(tk.END, f"Labels executed and saved as executed_labels.pdf\n")
        output_text.config(state=tk.DISABLED)
    else:
        output_text.config(state=tk.NORMAL)
        output_text.insert(tk.END, f"No labels and qrcodes to execute\n")
        output_text.config(state=tk.DISABLED)

# Create the main window
root = tk.Tk()
root.title("Label Generator")

# Create a frame for the input section
input_frame = tk.Frame(root)
input_frame.pack(side=tk.TOP, fill=tk.X)

# Create input labels and entry fields
product_code_label = tk.Label(input_frame, text="Product Code:")
product_code_label.pack(side=tk.LEFT)
product_code_entry = tk.Entry(input_frame)
product_code_entry.pack(side=tk.LEFT)

number_of_labels_label = tk.Label(input_frame, text="Number of Labels:")
number_of_labels_label.pack(side=tk.LEFT)
number_of_labels_entry = tk.Entry(input_frame, text="1")
number_of_labels_entry.pack(side=tk.LEFT)

# Create buttons in the input section
add_label_button = tk.Button(input_frame, text="Add Label", command=add_label)
add_label_button.pack(side=tk.LEFT)
clear_label_button = tk.Button(input_frame, text="Clear Labels", command=clear_labels)
clear_label_button.pack(side=tk.LEFT)
execute_label_button = tk.Button(input_frame, text="Execute Label", command=execute_label)
execute_label_button.pack(side=tk.LEFT)

# Create the output section with an increased height
output_text = tk.Text(root, height=20, width=100)
output_text.pack()

# Disable the output_text widget initially
output_text.config(state=tk.DISABLED)

# Start the GUI main loop
root.mainloop()