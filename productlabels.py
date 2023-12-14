from openpyxl import load_workbook
from reportlab.lib.pagesizes import portrait, A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
import qrcode

# Constants for label size and margins
LABEL_WIDTH = 2.2 * cm
LABEL_HEIGHT = 1.2 * cm
LABEL_MARGIN_LEFT = 0.5 * cm
LABEL_MARGIN_RIGHT = 0.5 * cm
LABEL_MARGIN_TOP = 1 * cm
LABEL_MARGIN_BOTTOM = 1 * cm
LABEL_MARGIN_X = 0.3 * cm
LABEL_MARGIN_Y = 0.2 * cm

# Constants for number of labels per row and column
LABELS_PER_ROW = 8
LABELS_PER_COLUMN = 20

# Constants for QR code and product code placement
QR_CODE_SIZE = 1 * cm
PRODUCT_CODE_FONT_SIZE = 4

# Load the Excel sheet
workbook = load_workbook('/Users/ersandagdeviren/Desktop/hwobarcode.xlsx')
sheet = workbook.active

# Create a PDF document
pdf = canvas.Canvas('/Users/ersandagdeviren/Desktop/labels.pdf', pagesize=portrait(A4))

# Set the initial position for labels
x = LABEL_MARGIN_LEFT
y = A4[1] - LABEL_MARGIN_TOP

# Counter for labels on the current page
label_counter = 0

# Iterate through the rows in the Excel sheet
for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
    product_code = row[0]
    product_quantity = row[2]

    # Generate QR code and labels based on quantity
    for label_count in range(product_quantity):
        # Check if we need to start a new page
        if label_counter >= LABELS_PER_ROW * LABELS_PER_COLUMN:
            pdf.showPage()
            x = LABEL_MARGIN_LEFT
            y = A4[1] - LABEL_MARGIN_TOP
            label_counter = 0

        # Generate QR code
        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(product_code)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_img_path = f'qr_code_{i}_{label_count+1}.png'  # Unique QR code image path for each label
        qr_img.save(qr_img_path)

        # Calculate the center position for the label
        label_center_x = x + (LABEL_WIDTH / 2)
        label_center_y = y - (LABEL_HEIGHT / 2)

        # Calculate the position for the QR code
        qr_x = label_center_x - (QR_CODE_SIZE / 2)
        qr_y = label_center_y - (QR_CODE_SIZE / 2)

        # Calculate the position for the product code
        product_code_x = label_center_x
        product_code_y = qr_y - 0.1 * cm  # Decreased gap between QR code and product code

        # Draw the label
        pdf.drawImage(qr_img_path, qr_x, qr_y, width=QR_CODE_SIZE, height=QR_CODE_SIZE)
        pdf.setFont("Helvetica", PRODUCT_CODE_FONT_SIZE)
        pdf.drawCentredString(product_code_x, product_code_y, str(product_code))

        # Move to the next label position
        if label_counter % LABELS_PER_ROW == LABELS_PER_ROW - 1:
            x = LABEL_MARGIN_LEFT
            y -= LABEL_HEIGHT + LABEL_MARGIN_Y
        else:
            x += LABEL_WIDTH + LABEL_MARGIN_X

        label_counter += 1

# Save the PDF file
pdf.save()